from datetime import timedelta
import json
import openpyxl

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.utils.timezone import localdate, now
from datetime import date
from django.db import models
from django.db.models import Count, Q
from django.forms import modelformset_factory
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView, DetailView
)

from .models import Integrante, Reunion, Asistencia
from .forms import IntegranteForm, ReunionForm, AsistenciaForm
from django.http import HttpResponse
from openpyxl import Workbook

from django.http import HttpResponse
from openpyxl.styles import Font

from django.db.models import Max
from django.views.decorators.http import require_POST


# Create your views here.
# CRUD Integrante

class IntegranteListView(LoginRequiredMixin, ListView):
    model = Integrante
    template_name = 'asistencia/integrante_list.html'
    context_object_name = 'integrantes'
    paginate_by = 25
    ordering = ['nombre']

    def get_queryset(self):
        queryset = Integrante.objects.all().order_by('nombre')

        q = self.request.GET.get('q')
        rol = self.request.GET.get('rol')
        tipo = self.request.GET.get('tipo')
        activo = self.request.GET.get('activo')
        sexo = self.request.GET.get('sexo')

        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) |
                Q(apellido__icontains=q)
            )

        if rol:
            queryset = queryset.filter(rol=rol)

        if tipo:
            queryset = queryset.filter(tipo=tipo)

        if activo in ['0', '1']:
            queryset = queryset.filter(activo=bool(int(activo)))
        
        if sexo: 
            queryset = queryset.filter(sexo=sexo)
        
        today = date.today()

        for integrante in queryset:
            if integrante.fecha_nacimiento:
                integrante.es_cumple = (
                    integrante.fecha_nacimiento.day == today.day and
                    integrante.fecha_nacimiento.month == today.month
                )
            else:
                integrante.es_cumple = False

        # ALERTA DE INASISTENCIA

        for integrante in queryset:
            # Mapear tipo de integrante a tipo de reunión
            tipo_reunion = 'N' if integrante.tipo == 'niño' else 'J'

            # Contar ausencias totales en reuniones de su tipo
            ausencias = Asistencia.objects.filter(
                integrante=integrante,
                reunion__tipo=tipo_reunion,
                presente=False
            ).count()

            integrante.alerta_inasistencia = ausencias >= 3 

        return queryset
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_resultados'] = self.get_queryset().count()
        return context

class IntegranteExportExcelView(LoginRequiredMixin, ListView):
    model = Integrante

    def get_queryset(self):
        qs = Integrante.objects.all()

        q = self.request.GET.get('q')
        rol = self.request.GET.get('rol')
        tipo = self.request.GET.get('tipo')
        activo = self.request.GET.get('activo')

        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(apellido__icontains=q)
            )

        if rol:
            qs = qs.filter(rol=rol)

        if tipo:
            qs = qs.filter(tipo=tipo)

        if activo in ['0', '1']:
            qs = qs.filter(activo=bool(int(activo)))

        return qs.order_by('nombre')
    
    def get(self, request, *args, **kwargs):
        integrantes = self.get_queryset()

        # Construir nombre dinámico
        nombre_archivo = "integrantes"

        rol = request.GET.get('rol')
        tipo = request.GET.get('tipo')
        activo = request.GET.get('activo')

        if rol:
            nombre_archivo += f"_{rol}"

        if tipo:
            nombre_archivo += f"_{tipo}"

        if activo == '1':
            nombre_archivo += "_activos"
        elif activo == '0':
            nombre_archivo += "_inactivos"

        nombre_archivo += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Integrantes"

        ws.append([
        "Nombre",
        "Apellido",
        "Rol",
        "Tipo",
        "Teléfono",
        "Email",
        "Activo"
    ])

        for i in integrantes:
            ws.append([
            i.nombre,
            i.apellido,
            i.rol,
            i.tipo,
            i.telefono,
            i.email,
            "Sí" if i.activo else "No"
        ])

        response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

        wb.save(response)
        return response

class IntegranteCreateView(LoginRequiredMixin, CreateView):
    model = Integrante
    form_class = IntegranteForm
    template_name = 'asistencia/integrante_form.html'
    success_url = reverse_lazy('integrante_list')

    def form_valid(self, form):
        messages.success(self.request, "Integrante guardado exitosamente.")
        return super().form_valid(form)

class IntegranteUpdateView(LoginRequiredMixin, UpdateView):
    model = Integrante
    form_class = IntegranteForm
    template_name = 'asistencia/integrante_form.html'
    success_url = reverse_lazy('integrante_list')

class IntegranteDeleteView(LoginRequiredMixin, DeleteView):
    model = Integrante
    template_name = 'asistencia/integrante_confirm_delete.html'
    success_url = reverse_lazy('integrante_list')


# Vista para marcar asistencia (funcional)
@login_required
def marcar_asistencia(request, reunion_id):
    reunion = get_object_or_404(Reunion, pk=reunion_id)

    integrantes = Integrante.objects.filter(activo=True).order_by('nombre')

    for integrante in integrantes:
        Asistencia.objects.get_or_create(
            reunion=reunion,
            integrante=integrante
        )

    queryset = Asistencia.objects.filter(
        reunion=reunion
    ).select_related('integrante').order_by('integrante__nombre')

    AsistenciaFormSet = modelformset_factory(
        Asistencia,
        form=AsistenciaForm,
        extra=0
    )

    if request.method == 'POST':
        formset = AsistenciaFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Asistencia registrada exitosamente.")
            return redirect('ver_asistencia', reunion_id=reunion.id)
    else:
        formset = AsistenciaFormSet(queryset=queryset)

    return render(request, 'asistencia/marcar_asistencia.html', {
        'reunion': reunion,
        'formset': formset
    })

# ---------- CRUD Reuniones ----------

class ReunionListView(LoginRequiredMixin, ListView):
    model = Reunion
    template_name = 'asistencia/reunion_list.html'
    context_object_name = 'reuniones'
    paginate_by = 25
    ordering = ['-fecha']

    def get_queryset(self):
        queryset = Reunion.objects.all().order_by('-fecha')

        q = self.request.GET.get('q')
        tipo = self.request.GET.get('tipo')
        maestro = self.request.GET.get('maestro')

        if q:
            queryset = queryset.filter(tema__icontains=q)

        if tipo:
            queryset = queryset.filter(tipo=tipo)

        if maestro:
            queryset = queryset.filter(maestro__id=maestro)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_resultados'] = self.get_queryset().count()
        context['maestros'] = Integrante.objects.filter(rol__in=['Maestro', 'Lider']).order_by('nombre')
        return context

class ReunionCreateView(LoginRequiredMixin, CreateView):
    model = Reunion
    form_class = ReunionForm
    template_name = 'asistencia/reunion_form.html'
    success_url = reverse_lazy('reunion_list')

    def form_valid(self, form):
        messages.success(self.request, "Reunión guardada exitosamente.")
        return super().form_valid(form)

class ReunionUpdateView(LoginRequiredMixin, UpdateView):
    model = Reunion
    form_class = ReunionForm
    template_name = 'asistencia/reunion_form.html'
    success_url = reverse_lazy('reunion_list')

class ReunionDeleteView(LoginRequiredMixin, DeleteView):
    model = Reunion
    template_name = 'asistencia/reunion_confirm_delete.html'
    success_url = reverse_lazy('reunion_list')

class ReunionDetailView(LoginRequiredMixin, DetailView):
    model = Reunion
    template_name = 'asistencia/reunion_detail.html'
    context_object_name = 'reunion'


@login_required
def dashboard(request):

    # ====== MÉTRICAS EXISTENTES ======
    total_integrantes = Integrante.objects.filter(activo=True).count()
    total_reuniones = Reunion.objects.count()

    ultima_reunion = Reunion.objects.order_by('-fecha').first()
    asistencia_ultima = 0

    if ultima_reunion:
        asistencia_ultima = Asistencia.objects.filter(
            reunion=ultima_reunion,
            presente=True
        ).count()

    fecha_limite = timezone.now().date() - timedelta(days=21)
    inactivos_21 = Integrante.objects.filter(
        activo=True,
        ultima_asistencia__lt=fecha_limite
    ).count()

    # Integrantes con alerta de inasistencia (3+ ausencias en reuniones de su tipo)
    integrantes_alerta = []
    for integrante in Integrante.objects.all():
        tipo_reunion = 'N' if integrante.tipo == 'niño' else 'J'
        ausencias = Asistencia.objects.filter(
            integrante=integrante,
            reunion__tipo=tipo_reunion,
            presente=False
        ).count()
        if ausencias >= 3:
            integrantes_alerta.append(integrante)

    # ====== CUMPLEAÑOS DE HOY ======
    hoy = date.today()

    cumple_hoy = Integrante.objects.filter(
        fecha_nacimiento__day=hoy.day,
        fecha_nacimiento__month=hoy.month,
        activo=True
    ).order_by('nombre')

    # ====== DATOS PARA CHART.JS ======
    reuniones = (
        Reunion.objects
        .annotate(
            asistentes=Count(
                'asistencia',
                filter=Q(asistencia__presente=True)
            )
        )
        .order_by('fecha')
    )

    chart_labels = [r.fecha.strftime('%d %b %Y') for r in reuniones]
    chart_data = [r.asistentes for r in reuniones]

    # ====== CONTEXTO ======
    context = {
        'total_integrantes': total_integrantes,
        'total_reuniones': total_reuniones,
        'asistencia_ultima': asistencia_ultima,
        'integrantes_alerta': integrantes_alerta,
        'ultima_reunion': ultima_reunion,

        # Cumpleaños
        'cumple_hoy': cumple_hoy,

        # Chart
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
    }

    return render(request, 'asistencia/dashboard.html', context)


@login_required
def ver_asistencia(request, reunion_id):
    reunion = get_object_or_404(Reunion, pk=reunion_id)
    asistencias = Asistencia.objects.filter(
        reunion=reunion
    ).select_related('integrante').order_by('integrante__nombre')

    return render(request, 'asistencia/ver_asistencia.html', {
        'reunion': reunion,
        'asistencias': asistencias
    })

def exportar_asistencia_excel(request, reunion_id):
    reunion = get_object_or_404(Reunion, id=reunion_id)

    asistencias = Asistencia.objects.filter(
        reunion=reunion,
        presente=True
    ).select_related('integrante')

    wb = openpyxl.Workbook()
    ws = wb.active

    fecha = reunion.fecha.strftime('%d de %B de %Y')
    ws.title = f"Asistencia {reunion.fecha.strftime('%d-%m-%Y')}"

    # TÍTULO
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Asistencia reunión del {fecha}"
    ws['A1'].font = Font(bold=True)

    # ENCABEZADOS
    ws['A3'] = 'Nombre'
    ws['B3'] = 'Comentario'

    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)

    fila = 4
    for a in asistencias:
        ws[f'A{fila}'] = f"{a.integrante.nombre} {a.integrante.apellido}"
        ws[f'B{fila}'] = a.comentario or ''
        fila += 1

    # AJUSTAR COLUMNAS
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    nombre_archivo = f"asistencia_reunion_{reunion.fecha.strftime('%d_%m_%Y')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

    wb.save(response)
    return response

def integrante_perfil(request, pk):
    integrante = get_object_or_404(Integrante, pk=pk)

    hoy = date.today()
    año_actual = hoy.year
    mes_actual = hoy.month

    asistencias = Asistencia.objects.filter(
        integrante=integrante
    ).select_related('reunion')

    # TOTAL HISTÓRICO
    total_asistencias = asistencias.filter(presente=True).count()

    # TOTAL DEL AÑO ACTUAL
    total_asistencias_año = asistencias.filter(
        presente=True,
        reunion__fecha__year=año_actual
    ).count()

    # TOTAL DEL MES ACTUAL
    total_asistencias_mes = asistencias.filter(
        presente=True,
        reunion__fecha__year=año_actual,
        reunion__fecha__month=mes_actual
    ).count()

    # HISTORIAL COMPLETO
    historial_asistencias = (
        asistencias
        .filter(presente=True)
        .order_by('-reunion__fecha')
    )

    # ÚLTIMA ASISTENCIA
    ultima_asistencia = historial_asistencias.first()

    context = {
        'integrante': integrante,
        'total_asistencias': total_asistencias,
        'total_asistencias_año': total_asistencias_año,
        'total_asistencias_mes': total_asistencias_mes,
        'historial_asistencias': historial_asistencias,
        'ultima_asistencia': ultima_asistencia,
    }

    return render(request, 'asistencia/integrante_perfil.html', context)

@require_POST
def integrante_cambiar_foto(request, pk):
    integrante = get_object_or_404(Integrante, pk=pk)

    if 'foto' in request.FILES:
        integrante.foto = request.FILES['foto']
        integrante.save()

    return redirect('integrante_perfil', pk=pk)

@require_POST
def integrante_editar_descripcion(request, pk):
    integrante = get_object_or_404(Integrante, pk=pk)
    descripcion = request.POST.get('descripcion', '').strip()

    integrante.descripcion = descripcion
    integrante.save()

    return redirect('integrante_perfil', pk=pk)

@require_POST
def integrante_toggle_activo(request, pk):
    integrante = get_object_or_404(Integrante, pk=pk)
    integrante.activo = not integrante.activo
    integrante.save()
    return redirect('integrante_perfil', pk=pk)



